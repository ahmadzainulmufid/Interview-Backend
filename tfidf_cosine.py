from collections import Counter
import math
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory

stemmer = StemmerFactory().create_stemmer()

STOPWORDS_ID = {
    "yang", "untuk", "pada", "ke", "para", "namun", "menurut", "antara", "dia",
    "dua", "ini", "itu", "dan", "atau", "juga", "sehingga", "sebagai",
    "dari", "di", "ada", "adalah", "akan", "bahwa", "dalam", "dengan",
    "karena", "oleh", "saat", "seperti", "yaitu", "yakni", "sebuah",
    "suatu", "adapun", "agar", "bagi", "begitu", "belum", "boleh",
    "hal", "hingga", "jika", "kali", "kata", "lagi", "lain", "lalu",
    "maka", "masih", "mereka", "namun", "pun", "sama", "sangat",
    "sedang", "sejak", "sementara", "serta", "sesuatu", "setelah",
    "sudah", "supaya", "tanpa", "tapi", "telah", "terhadap", "tetapi",
    "tidak", "tersebut", "yg", "the", "a", "an", "of", "is", "are",
    "apa", "bagaimana", "baik", "demikian", "cara", "harus", "merupakan",
    "satu", "dua", "tiga", "empat", "lima", "enam", "tujuh",
    "delapan", "sembilan", "sepuluh", "pertama", "kedua", "ketiga",
}

SYNONYM_MAP = {
    "chap": "cap",
    "ketersediaan": "availability",
    "konsistensi": "consistency",
    "toleransi partisi": "partition tolerance",
    "layanan mikro": "microservice",
    "microservices": "microservice",
    "basis data": "database",
    "antrian pesan": "message queue",
    "penyeimbang beban": "load balancer",
}

TOKEN_PATTERN = re.compile(r"[a-z]{3,}")


TITLE_FONT = Font(bold=True, size=13, color="FFFFFF")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
LABEL_FONT = Font(bold=True, size=10)
BODY_FONT = Font(size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")


def normalize_synonyms(text: str) -> str:
    lowered = text.lower()
    for source, target in SYNONYM_MAP.items():
        lowered = re.sub(rf"\b{re.escape(source)}\b", target, lowered)
    return lowered


def case_folding(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokenizing(text: str) -> list[str]:
    return [t for t in text.split(" ") if t]


def filtering(tokens: list[str]) -> list[str]:
    return [t for t in tokens if t not in STOPWORDS_ID]


_stem_cache = {}


def stemming(tokens: list[str]) -> list[str]:
    hasil = []
    for token in tokens:
        if token not in _stem_cache:
            _stem_cache[token] = stemmer.stem(token)
        hasil.append(_stem_cache[token])
    return hasil


def preprocess_text(text: str) -> list[str]:
    normalized = normalize_synonyms(text)
    folded = case_folding(normalized)
    tokens = tokenizing(folded)
    tokens = filtering(tokens)
    tokens = stemming(tokens)
    tokens = [t for t in tokens if t and t not in STOPWORDS_ID]
    return tokens


def compute_tf(tokens: list[str]) -> dict[str, float]:
    """
    Menghitung TF ternormalisasi:
    jumlah kemunculan term dibagi jumlah seluruh term dalam dokumen.
    """
    if not tokens:
        return {}

    counts = Counter(tokens)
    total_terms = len(tokens)

    return {
        term: count / total_terms
        for term, count in counts.items()
    }


def compute_idf_local(corpus_tokens: list[list[str]]):
    """
    Hitung IDF LOKAL hanya dari dokumen yang sedang dibandingkan
    (D1 & D2), TANPA corpus eksternal / knowledge base global.

    Rumus: IDF = ln(N / DF) + 1
    """
    N = len(corpus_tokens)
    df = Counter()
    for tokens in corpus_tokens:
        for term in set(tokens):
            df[term] += 1
    idf_dict = {}
    for term, freq in df.items():
        idf_dict[term] = math.log(N / freq) + 1.0
    # default_idf dipakai untuk term yang (secara teori) DF=0 / tak dikenal.
    # Karena vocab selalu diambil dari gabungan D1 & D2, DF minimal = 1,
    # sehingga nilai ini praktis tidak pernah dipakai.
    default_idf = math.log(N / 1) + 1.0
    return idf_dict, df, N, default_idf


def compute_tfidf_vector(tf: dict, idf: dict, vocab: list[str], default_idf: float) -> list[float]:
    return [tf.get(t, 0.0) * idf.get(t, default_idf) for t in vocab]


def calculate_cosine_similarity(vec1, vec2) -> float:
    dot = sum(a * b for a, b in zip(vec1, vec2))
    mag1 = math.sqrt(sum(a ** 2 for a in vec1))
    mag2 = math.sqrt(sum(b ** 2 for b in vec2))
    if mag1 == 0 or mag2 == 0:
        return 0.0
    return dot / (mag1 * mag2)

def tabel_ti(list_d1: list[str], list_d2: list[str]) -> pd.DataFrame:
    """Gabungkan token D1 & D2 sejajar per-index (ti) untuk ditampilkan."""
    n = max(len(list_d1), len(list_d2))
    rows = []
    for i in range(n):
        rows.append({
            "ti": f"t{i + 1}",
            "Token D1": list_d1[i] if i < len(list_d1) else "",
            "Token D2": list_d2[i] if i < len(list_d2) else "",
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=["ti", "Token D1", "Token D2"])


def cetak_tabel(df: pd.DataFrame, judul: str):
    print(f"\n=== {judul} ===")
    if df.empty:
        print("(kosong)")
    else:
        print(df.to_string(index=False))


def _write_df_to_sheet(ws, df: pd.DataFrame):
    """Tulis DataFrame ke worksheet openpyxl dengan header berwarna."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=str(col_name))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()]
        ) if not df.empty else len(str(col_name))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def export_to_excel(path: str, data: dict):
    wb = Workbook()

    ws = wb.active
    ws.title = "Ringkasan"
    ws["A1"] = "RINGKASAN ANALISIS COSINE SIMILARITY (IDF Lokal, N=%d)" % data["N_local"]
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws.merge_cells("A1:B1")

    ringkasan_rows = [
        ("Pertanyaan", data["pertanyaan"]),
        ("Jawaban Referensi (D1)", data["jawaban_referensi"]),
        ("Jawaban Kandidat (D2)", data["jawaban_kandidat"]),
        ("Ukuran Corpus IDF (N)", data["N_local"]),
        ("Dot Product (D1 . D2)", round(data["dot_product"], 6)),
        ("Magnitude ||D1||", round(data["mag_d1"], 6)),
        ("Magnitude ||D2||", round(data["mag_d2"], 6)),
        ("Cosine Similarity", round(data["cosine"], 6)),
        ("Skor Akhir ", round(data["skor_akhir"], 4)),
    ]
    r = 3
    for label, val in ringkasan_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    ws_cf = wb.create_sheet("1_CaseFolding")
    ws_cf["A1"] = "Dokumen"
    ws_cf["B1"] = "Hasil Case Folding"
    ws_cf["A1"].font = HEADER_FONT
    ws_cf["B1"].font = HEADER_FONT
    ws_cf["A1"].fill = HEADER_FILL
    ws_cf["B1"].fill = HEADER_FILL
    ws_cf["A2"] = "D1 (Referensi)"
    ws_cf["B2"] = data["cf_d1"]
    ws_cf["A3"] = "D2 (Kandidat)"
    ws_cf["B3"] = data["cf_d2"]
    for row in (2, 3):
        ws_cf.cell(row=row, column=1).font = BODY_FONT
        c = ws_cf.cell(row=row, column=2)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws_cf.column_dimensions["A"].width = 20
    ws_cf.column_dimensions["B"].width = 100

    sheet_specs = [
        ("2_Tokenizing", data["df_tokenizing"]),
        ("3_Filtering", data["df_filtering"]),
        ("4_Stemming", data["df_stemming"]),
        ("5_Vocabulary", data["df_vocab"]),
        ("6_TF", data["df_tf"]),
        ("7_DF", data["df_df"]),
        ("8_IDF", data["df_idf"]),
        ("9_TFIDF", data["df_tfidf"]),
    ]
    for name, df in sheet_specs:
        ws2 = wb.create_sheet(name)
        _write_df_to_sheet(ws2, df)

    ws_cos = wb.create_sheet("10_CosineSimilarity")
    ws_cos["A1"] = "Metrik"
    ws_cos["B1"] = "Nilai"
    ws_cos["A1"].font = HEADER_FONT
    ws_cos["B1"].font = HEADER_FONT
    ws_cos["A1"].fill = HEADER_FILL
    ws_cos["B1"].fill = HEADER_FILL
    cos_rows = [
        ("Dot Product (D1 . D2)", round(data["dot_product"], 6)),
        ("Magnitude ||D1||", round(data["mag_d1"], 6)),
        ("Magnitude ||D2||", round(data["mag_d2"], 6)),
        ("Cosine Similarity", round(data["cosine"], 6)),
        ("Skor Akhir ", round(data["skor_akhir"], 4)),
    ]
    for i, (k, v) in enumerate(cos_rows, start=2):
        ws_cos.cell(row=i, column=1, value=k).font = BODY_FONT
        ws_cos.cell(row=i, column=2, value=v).font = BODY_FONT
    ws_cos.column_dimensions["A"].width = 28
    ws_cos.column_dimensions["B"].width = 18

    wb.save(path)


pertanyaan = (
    "Apa yang membuat antarmuka (interface) menjadi komponen penting dalam "
    "pemrograman, dan bagaimana hal ini memengaruhi kemudahan perawatan serta "
    "pengembangan sistem yang kompleks?"
)

jawaban_referensi = (
    "Antarmuka penting dalam pemrograman karena memungkinkan abstraksi dan "
    "modularitas, memisahkan antarmuka dari implementasi, sehingga meningkatkan "
    "kemudahan dan kemudahan perawatan kode. Antarmuka juga membantu mengurangi "
    "ketergantungan antar komponen dan meningkatkan penggunaan kembali kode. "
    "Dengan demikian, antarmuka mempermudah pengembangan dan pemeliharaan sistem "
    "yang kompleks."
)

jawaban_kandidat = (
    "Baik, untuk antarmuka merupakan sebuah proses yang sangat penting dalam pemrograman."
    "Antarmuka juga membantu mengurangi ketergantungan antar komponen dan meningkatkan penggunaan kembali kode."
)


def main():
    print("#" * 100)
    print("PERTANYAAN :", pertanyaan)
    print("-" * 100)
    print("JAWABAN REFERENSI (D1):", jawaban_referensi)
    print("-" * 100)
    print("JAWABAN KANDIDAT  (D2):", jawaban_kandidat)
    print("#" * 100)

    # TAHAP 1: CASE FOLDING
    cf_d1 = case_folding(normalize_synonyms(jawaban_referensi))
    cf_d2 = case_folding(normalize_synonyms(jawaban_kandidat))
    print("\n=== TAHAP 1: CASE FOLDING ===")
    print("D1:", cf_d1)
    print("D2:", cf_d2)

    # TAHAP 2: TOKENIZING
    tok_d1 = tokenizing(cf_d1)
    tok_d2 = tokenizing(cf_d2)
    df_tokenizing = tabel_ti(tok_d1, tok_d2)
    cetak_tabel(df_tokenizing, "TAHAP 2: TOKENIZING")

    # TAHAP 3: FILTERING
    filt_d1 = filtering(tok_d1)
    filt_d2 = filtering(tok_d2)
    df_filtering = tabel_ti(filt_d1, filt_d2)
    cetak_tabel(df_filtering, "TAHAP 3: FILTERING (Stopword Removal)")

    # TAHAP 4: STEMMING
    stem_d1 = stemming(filt_d1)
    stem_d2 = stemming(filt_d2)
    stem_d1 = [t for t in stem_d1 if t and t not in STOPWORDS_ID]
    stem_d2 = [t for t in stem_d2 if t and t not in STOPWORDS_ID]
    df_stemming = tabel_ti(stem_d1, stem_d2)
    cetak_tabel(df_stemming, "TAHAP 4: STEMMING (hasil akhir preprocessing)")

    # VOCAB — hanya term yang muncul di D1/D2
    vocab = sorted(set(stem_d1) | set(stem_d2))
    df_vocab = pd.DataFrame({"ti": [f"t{i + 1}" for i in range(len(vocab))], "Term": vocab})
    cetak_tabel(df_vocab, "DAFTAR TERM HASIL PREPROCESSING (Vocabulary Gabungan D1 & D2)")

    # TAHAP 5: TF (per-dokumen)
    tf_d1 = compute_tf(stem_d1)
    tf_d2 = compute_tf(stem_d2)
    df_tf = pd.DataFrame({
        "Term": vocab,
        "TF D1": [round(tf_d1.get(t, 0.0), 4) for t in vocab],
        "TF D2": [round(tf_d2.get(t, 0.0), 4) for t in vocab],
    })
    cetak_tabel(df_tf, "TAHAP 5: TERM FREQUENCY (TF)")

    # ── IDF LOKAL: dihitung HANYA dari D1 & D2 (N=2), TANPA corpus global ──
    idf_local, df_local, N_local, default_idf = compute_idf_local([stem_d1, stem_d2])

    # TAHAP 6: DF (lokal, dari D1 & D2 saja)
    df_df = pd.DataFrame({
        "Term": vocab,
        "DF (D1 & D2)": [df_local.get(t, 0) for t in vocab],
    })
    cetak_tabel(df_df, "TAHAP 6: DOCUMENT FREQUENCY (DF) — lokal dari D1 & D2 (N=2)")

    # TAHAP 7: IDF (lokal)
    df_idf = pd.DataFrame({
        "Term": vocab,
        "DF": [df_local.get(t, 0) for t in vocab],
        "N (lokal)": [N_local for _ in vocab],
        "IDF = ln(N/DF)+1": [round(idf_local.get(t, default_idf), 4) for t in vocab],
    })
    cetak_tabel(df_idf, f"TAHAP 7: INVERSE DOCUMENT FREQUENCY (IDF) — lokal, N={N_local}")

    # TAHAP 8: TF-IDF
    tfidf_d1 = compute_tfidf_vector(tf_d1, idf_local, vocab, default_idf)
    tfidf_d2 = compute_tfidf_vector(tf_d2, idf_local, vocab, default_idf)
    df_tfidf = pd.DataFrame({
        "Term": vocab,
        "TF D1": [round(tf_d1.get(t, 0.0), 4) for t in vocab],
        "TF D2": [round(tf_d2.get(t, 0.0), 4) for t in vocab],
        "IDF (lokal)": [round(idf_local.get(t, default_idf), 4) for t in vocab],
        "TF-IDF D1": [round(w, 4) for w in tfidf_d1],
        "TF-IDF D2": [round(w, 4) for w in tfidf_d2],
    })
    cetak_tabel(df_tfidf, "TAHAP 8: PEMBOBOTAN TF-IDF (W = TF x IDF_lokal)")

    # TAHAP 9: COSINE SIMILARITY
    dot_product = sum(a * b for a, b in zip(tfidf_d1, tfidf_d2))
    mag_d1 = math.sqrt(sum(a ** 2 for a in tfidf_d1))
    mag_d2 = math.sqrt(sum(b ** 2 for b in tfidf_d2))
    cosine = calculate_cosine_similarity(tfidf_d1, tfidf_d2)
    skor_akhir = round(cosine, 4)

    print("\n=== TAHAP 9: COSINE SIMILARITY ===")
    print(f"Dot Product (D1 . D2)      : {dot_product:.6f}")
    print(f"Magnitude ||D1||           : {mag_d1:.6f}")
    print(f"Magnitude ||D2||           : {mag_d2:.6f}")
    print(f"Cosine Similarity          : {cosine:.6f}")
    print(f"Skor Akhir                 : {skor_akhir}")

    output_path = "hasil_analisis_cosine_similarity.xlsx"
    export_to_excel(output_path, {
        "pertanyaan": pertanyaan,
        "jawaban_referensi": jawaban_referensi,
        "jawaban_kandidat": jawaban_kandidat,
        "cf_d1": cf_d1,
        "cf_d2": cf_d2,
        "df_tokenizing": df_tokenizing,
        "df_filtering": df_filtering,
        "df_stemming": df_stemming,
        "df_vocab": df_vocab,
        "df_tf": df_tf,
        "df_df": df_df,
        "df_idf": df_idf,
        "df_tfidf": df_tfidf,
        "dot_product": dot_product,
        "mag_d1": mag_d1,
        "mag_d2": mag_d2,
        "cosine": cosine,
        "skor_akhir": skor_akhir,
        "N_local": N_local,
    })
    print(f"\ntabel berhasil diekspor otomatis ke: {output_path}")

    return {
        "vocab": vocab,
        "tf_d1": tf_d1,
        "tf_d2": tf_d2,
        "df": df_local,
        "idf": idf_local,
        "tfidf_d1": tfidf_d1,
        "tfidf_d2": tfidf_d2,
        "cosine_similarity": cosine,
        "skor_akhir": skor_akhir,
    }


if __name__ == "__main__":
    main()