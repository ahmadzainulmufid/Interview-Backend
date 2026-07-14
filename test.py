import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("human_eval_results.json", encoding="utf-8") as f:
    raw = json.load(f)

FEEDBACK = [
    "Jawaban Anda sudah mencakup penjelasan bahwa interpretasi mengeksekusi kode baris demi baris tanpa menghasilkan file executable terpisah, dan kompilasi menghasilkan file mesin yang bisa dieksekusi berulang. Untuk memperkuat jawaban, tambahkan implikasi praktis dari kedua pendekatan, misalnya soal kecepatan eksekusi dan portabilitas platform.",
    "Jawaban Anda sudah mencakup dua dari lima prinsip SOLID (Single Responsibility dan Dependency Inversion), namun belum membahas tiga prinsip lainnya yaitu Open/Closed, Liskov Substitution, dan Interface Segregation. Untuk memperkuat jawaban, sebutkan seluruh lima prinsip beserta definisi singkatnya agar cakupan jawaban lebih lengkap.",
    "Jawaban Anda sudah mencakup konsep struktur node pada linked list, namun belum membahas keunggulan penyisipan/penghapusan data pada linked list dibanding array. Untuk memperkuat jawaban, jelaskan juga bagaimana array menyimpan elemen secara berdekatan di memori sehingga akses berbasis indeks menjadi lebih cepat.",
    "Jawaban Anda belum membahas bahwa Git merupakan sistem kontrol versi terdistribusi yang melacak perubahan kode serta mendukung manajemen cabang dan kerja kolaboratif. Untuk memperkuat jawaban, jelaskan bagaimana Git menyimpan riwayat perubahan (commit) dan memungkinkan banyak developer bekerja pada branch terpisah sebelum digabungkan.",
    "Jawaban Anda sudah mencakup gagasan bahwa arsitektur microservice memecah aplikasi menjadi layanan-layanan kecil yang berkomunikasi lewat API, namun belum membahas independensi tanggung jawab tiap layanan atas fungsionalitas tertentu. Untuk memperkuat jawaban, jelaskan bagaimana setiap service bisa dikembangkan dan di-deploy secara independen tanpa memengaruhi service lain.",
    "Jawaban Anda sudah mencakup perbedaan struktur (terurut vs key-value) dan kecepatan akses antara List dan HashMap, namun belum menegaskan secara eksplisit bahwa keunggulan HashMap adalah pada pencarian data spesifik berdasarkan kunci. Untuk memperkuat jawaban, tambahkan contoh kasus penggunaan konkret, misalnya kapan sebaiknya memilih HashMap dibanding List.",
    "Jawaban Anda belum membahas perbedaan mendasar antara tuple dan antarmuka dari sisi keamanan tipe data dan penamaan properti. Untuk memperkuat jawaban, jelaskan bahwa interface mendefinisikan kontrak dengan nama properti dan tipe data yang jelas serta lebih aman, sedangkan tuple lebih fleksibel namun tidak memiliki nama properti.",
    "Jawaban Anda sudah mencakup prinsip FIFO pada antrian, namun belum membahas bagaimana hashmap menyimpan elemen dalam bentuk pasangan kunci-nilai untuk akses cepat. Untuk memperkuat jawaban, jelaskan juga elemen pertama yang keluar pada antrian dan bagaimana ini berbeda dari cara kerja hashmap yang tidak memiliki urutan.",
    "Jawaban Anda belum membahas peran manajer paket dalam proses instalasi dan resolusi ketergantungan pustaka menggunakan istilah yang konsisten dengan konteks jawaban referensi. Untuk memperkuat jawaban, gunakan istilah seperti 'manajer paket', 'instalasi', dan 'resolusi ketergantungan' selain istilah teknis berbahasa Inggris yang sudah disebutkan.",
    "Jawaban Anda sudah menyinggung soal kerja kolaboratif pada sistem terdistribusi, namun belum membahas keuntungan konkret seperti kemampuan bekerja secara offline, redundansi data, dan fleksibilitas percabangan/penggabungan. Untuk memperkuat jawaban, sebutkan minimal dua keuntungan teknis spesifik dari sistem kontrol versi terdistribusi dibanding sistem tersentralisasi.",
    "Jawaban Anda sudah mencakup definisi dasar layanan stateless yang tidak menyimpan data antar request, namun belum membahas bagaimana layanan stateful menjaga status klien untuk kebutuhan sesi dan transaksi. Untuk memperkuat jawaban, berikan contoh konkret kapan masing-masing pendekatan lebih sesuai digunakan, misalnya API publik vs sistem checkout e-commerce.",
    "Jawaban Anda sudah mencakup definisi rekursi dan enkapsulasi beserta cara kerjanya secara akurat, namun belum menegaskan secara eksplisit bahwa enkapsulasi membungkus data dan metode dalam satu unit program yang sama. Untuk memperkuat jawaban, tambahkan contoh penerapan nyata dari masing-masing konsep dalam kode.",
    "Jawaban Anda belum membahas mekanisme spesifik pencapaian polimorfisme seperti method overriding dan method overloading. Untuk memperkuat jawaban, jelaskan bagaimana kedua mekanisme tersebut memungkinkan objek menangani berbagai jenis input atau situasi secara berbeda.",
    "Jawaban Anda sudah mencakup contoh notasi seperti O(n) dan O(log n), namun belum membahas bahwa Big O menggambarkan batas atas kompleksitas waktu algoritma untuk keperluan perbandingan efisiensi. Untuk memperkuat jawaban, jelaskan mengapa memahami tingkat pertumbuhan kompleksitas ini penting saat memilih algoritma untuk skala data besar.",
    "Jawaban Anda sudah mencakup penjelasan access modifier seperti private, protected, dan public dalam menyembunyikan detail implementasi, namun belum menegaskan secara eksplisit bagaimana access modifier ini mengontrol tampilan antarmuka yang diberikan kepada pengguna objek. Untuk memperkuat jawaban, berikan contoh kode singkat yang menunjukkan penerapan access modifier tersebut.",
]

IDEAL_ANSWERS = [
    "Kompilasi menerjemahkan kode sumber ke dalam kode mesin sehingga membuat file yang dapat dieksekusi. Interpretasi menerjemahkan dan mengeksekusi kode baris demi baris tanpa file yang dapat dieksekusi.",
    "SOLID mewakili lima prinsip desain untuk OOP: Tanggung Jawab Tunggal, Terbuka/Tertutup, Substitusi Liskov, Segregasi Antarmuka, Inversi Ketergantungan.",
    "Sebuah array memiliki ukuran tetap dan menyimpan elemen dalam memori yang berdekatan; daftar tertaut terdiri dari node dengan data dan referensi, memudahkan penyisipan dan penghapusan.",
    "Git adalah sistem kontrol versi terdistribusi untuk melacak perubahan kode sumber, memungkinkan kerja kolaboratif dan manajemen cabang.",
    "Arsitektur layanan mikro terdiri dari layanan kecil dan independen yang berkomunikasi melalui API, masing-masing bertanggung jawab atas fungsionalitas tertentu.",
    "List dan HashMap adalah dua struktur data yang berbeda. List adalah koleksi data yang terurut dan dapat diakses menggunakan indeks, sedangkan HashMap adalah koleksi data yang tidak terurut dan diakses menggunakan kunci unik. HashMap memiliki kecepatan akses yang lebih cepat untuk pencarian data spesifik, sedangkan List lebih cocok untuk urutan operasi dan iterasi.",
    "Tuple adalah kumpulan nilai yang dapat memiliki tipe data berbeda, sedangkan antarmuka adalah kontrak yang menentukan struktur dan tipe data yang harus diimplementasikan. Tuple lebih fleksibel, tetapi kurang aman dan tidak memiliki nama properti, sedangkan interface lebih aman dan memiliki nama properti yang jelas.",
    "Hashmap adalah struktur data yang menyimpan data dalam bentuk pasangan kunci-nilai, memungkinkan akses cepat ke nilai berdasarkan kunci. Sementara itu, queue adalah struktur data yang mengikuti prinsip FIFO, di mana elemen pertama yang dimasukkan akan menjadi elemen pertama yang dikeluarkan.",
    "Manajer paket mengelola instalasi, pembaruan, dan resolusi ketergantungan perpustakaan, menyederhanakan manajemen perpustakaan dalam pengembangan.",
    "Sistem terdistribusi memungkinkan pekerjaan offline, percabangan/penggabungan yang fleksibel, operasi yang lebih cepat, redundansi, dan alur kerja kolaboratif.",
    "Layanan tanpa kewarganegaraan tidak menyimpan data klien di antara permintaan; layanan stateful menjaga status klien, berguna untuk sesi dan transaksi.",
    "Rekursi adalah teknik pemrograman di mana fungsi memanggil dirinya sendiri, sedangkan enkapsulasi adalah konsep pemrograman di mana data dan metode dibungkus dalam satu unit, seperti kelas. Enkapsulasi menyembunyikan detail implementasi dan hanya menampilkan antarmuka yang diperlukan.",
    "Polimorfisme adalah kemampuan suatu objek untuk memiliki banyak bentuk atau perilaku, tergantung pada konteksnya. Hal ini dicapai melalui metode overriding atau metode overloading, yang memungkinkan objek untuk menangani berbagai jenis input atau situasi.",
    "Notasi Big O menggambarkan batas atas kompleksitas waktu algoritma, penting untuk membandingkan efisiensi dan tingkat pertumbuhan.",
    "Enkapsulasi bekerja dengan menyembunyikan detail implementasi internal dari sebuah objek dan hanya menampilkan antarmuka yang diperlukan kepada pengguna. Ini dilakukan dengan menggunakan pengubah akses seperti private, protected, dan public untuk mengontrol akses ke data dan metode.",
]

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
GOOD_FILL = PatternFill("solid", start_color="C6E0B4", end_color="C6E0B4")
AVG_FILL = PatternFill("solid", start_color="FFE699", end_color="FFE699")
BAD_FILL = PatternFill("solid", start_color="F8CBAD", end_color="F8CBAD")

wb = Workbook()
ws = wb.active
ws.title = "Human Evaluation"

ws["A1"] = "LEMBAR HUMAN EVALUATION - SISTEM PENILAIAN JAWABAN INTERVIEW (TF-IDF COSINE SIMILARITY)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:H1")

ws["A2"] = "Petunjuk: Nilai apakah Skor Sistem, Status Sistem, dan Feedback AI di bawah ini SESUAI dengan penilaian Anda sebagai pakar terhadap jawaban kandidat. Isi kolom G dengan VALID atau INVALID, dan kolom H dengan catatan bila perlu."
ws["A2"].font = Font(name="Arial", italic=True, size=10)
ws.merge_cells("A2:H2")

headers = ["No", "Pertanyaan", "Jawaban Referensi (Ideal)", "Jawaban Kandidat",
           "Skor Sistem (%)", "Status Sistem", "Feedback AI (Sistem)", "Penilaian Pakar (VALID/INVALID)"]

header_row = 4
for j, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

status_fill_map = {"Good": GOOD_FILL, "Average": AVG_FILL, "Bad": BAD_FILL}

for i, (item, ideal, feedback) in enumerate(zip(raw, IDEAL_ANSWERS, FEEDBACK), start=1):
    q, kandidat, score, status, found, missing = item
    row = header_row + i
    values = [i, q, ideal, kandidat, score, status, feedback, ""]
    for j, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=j, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if j in (1, 5, 6):
            cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        if j == 6:
            cell.fill = status_fill_map.get(status, None)

widths = {1: 5, 2: 32, 3: 40, 4: 40, 5: 12, 6: 12, 7: 45, 8: 20}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
ws.row_dimensions[header_row].height = 30

# Sheet ringkasan
ws2 = wb.create_sheet("Ringkasan Distribusi")
ws2["A1"] = "RINGKASAN DISTRIBUSI SAMPEL"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:B1")

from collections import Counter
status_counts = Counter(item[3] for item in raw)
r = 3
ws2.cell(row=r, column=1, value="Status Sistem").font = HEADER_FONT
ws2.cell(row=r, column=1).fill = HEADER_FILL
ws2.cell(row=r, column=2, value="Jumlah Sampel").font = HEADER_FONT
ws2.cell(row=r, column=2).fill = HEADER_FILL
for status in ["Good", "Average", "Bad"]:
    r += 1
    ws2.cell(row=r, column=1, value=status).font = BODY_FONT
    ws2.cell(row=r, column=2, value=status_counts.get(status, 0)).font = BODY_FONT
r += 2
ws2.cell(row=r, column=1, value="Total Sampel").font = Font(bold=True)
ws2.cell(row=r, column=2, value=len(raw)).font = Font(bold=True)
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 18

wb.save("human_evaluation_interview.xlsx")
print("Selesai. Distribusi status:", dict(status_counts))